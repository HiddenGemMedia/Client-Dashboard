from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_ROOT / "data" / "Client Performance Tracking (3).xlsx"
OUTPUT_PATH = PROJECT_ROOT / "data" / "performance-dashboard.json"
EXCLUDED_SHEETS = {
    "Overview",
    "Template",
    "Copy of Template",
    "Copy of Template 1",
}


def slugify(value: str) -> str:
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", value.lower()))


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = (
        text.replace("🎉", "")
        .replace("📈", "")
        .replace("👁️", "")
        .replace("/", " ")
        .replace("%", " pct ")
        .replace("#", " num ")
        .replace("$", " ")
    )
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def clean_number(value: object):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return None

    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace(",", "").replace("$", "").replace("%", "")
    try:
        number = float(text)
        return -number if negative else number
    except ValueError:
        return None


def percent_text(value: object):
    if value is None or value == "":
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None

    number = clean_number(value)
    if number is None:
        return None
    if abs(number) <= 1:
        return f"{round(number * 100)}%"
    return f"{round(number)}%"


def get_value(row_map: dict[str, object], *keys: str):
    for key in keys:
        if key in row_map:
            return row_map[key]
    return None


def build_row(row_map: dict[str, object]):
    timeline = get_value(row_map, "timeline")
    if not isinstance(timeline, datetime):
        return None

    return {
        "year": timeline.year,
        "month": timeline.month,
        "total_views": clean_number(get_value(row_map, "total_views")) or 0,
        "total_view_growth": percent_text(get_value(row_map, "total_view_growth")),
        "ig_views": clean_number(get_value(row_map, "ig_views")) or 0,
        "fb_views": clean_number(get_value(row_map, "fb_views")) or 0,
        "tiktok_views": clean_number(get_value(row_map, "tiktok_views")) or 0,
        "ig_followers": clean_number(get_value(row_map, "ig_followers")) or 0,
        "fb_followers": clean_number(get_value(row_map, "fb_followers")) or 0,
        "tiktok_followers": clean_number(get_value(row_map, "tiktok_followers")) or 0,
        "ttl_followers": clean_number(get_value(row_map, "ttl_followers")) or 0,
        "follower_growth_pct": clean_number(get_value(row_map, "follower_growth")) or 0,
        "website_traffic": clean_number(get_value(row_map, "website_traffic")) or 0,
        "ad_spend": clean_number(get_value(row_map, "ad_spend")) or 0,
        "cost_per_follower": clean_number(get_value(row_map, "cost_per_follower")),
        "cost_per_lead": clean_number(get_value(row_map, "cost_per_lead")),
        "cost_per_booking": clean_number(get_value(row_map, "cost_per_booking")),
        "new_leads": clean_number(get_value(row_map, "new_leads")) or 0,
        "ttl_leads": clean_number(get_value(row_map, "ttl_leads")) or 0,
        "lead_growth_pct": clean_number(get_value(row_map, "lead_growth")) or 0,
        "total_booking_revenue": clean_number(get_value(row_map, "total_booking_revenue")) or 0,
        "direct_booking_revenue": clean_number(get_value(row_map, "direct_booking_revenue")) or 0,
        "direct_booking_split_pct": clean_number(get_value(row_map, "direct_booking_split")) or 0,
        "ly_total_booking_revenue": clean_number(get_value(row_map, "ly_total_booking_revenue")) or 0,
        "ly_direct_booking_revenue": clean_number(get_value(row_map, "ly_direct_booking_revenue")) or 0,
        "ly_direct_booking_split_pct": clean_number(get_value(row_map, "ly_direct_booking_split")) or 0,
        "notes": str(get_value(row_map, "notes_insights") or "").strip(),
    }


def export_workbook():
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    payload = {"clients": [], "rowsByClientSlug": {}}

    for sheet_name in workbook.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            continue

        sheet = workbook[sheet_name]
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [normalize_header(value) for value in header_cells]
        rows = []

        for raw_row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value not in (None, "") for value in raw_row):
                continue
            row_map = {headers[index]: raw_row[index] for index in range(min(len(headers), len(raw_row)))}
            normalized = build_row(row_map)
            if normalized:
                rows.append(normalized)

        if not rows:
            continue

        slug = slugify(sheet_name)
        rows.sort(key=lambda row: (row["year"], row["month"]))
        payload["clients"].append({"slug": slug, "name": sheet_name})
        payload["rowsByClientSlug"][slug] = rows

    payload["clients"].sort(key=lambda client: client["name"].lower())
    OUTPUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    export_workbook()
