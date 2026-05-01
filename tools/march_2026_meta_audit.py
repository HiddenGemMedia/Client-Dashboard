from __future__ import annotations

import difflib
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path("/Users/Pooja/Documents/New project copy/inspired-retreats-dashboard")
WORKBOOK_PATH = PROJECT_ROOT / "Data" / "Ads ROI - Sheet.xlsx"
PERFORMANCE_JSON_PATH = PROJECT_ROOT / "Data" / "performance-dashboard.json"
SHEET_COLUMN_MAPS = {
    "March 2026": {
        "campaign_spend": 5,
        "impressions": 6,
        "profile_visits": 7,
        "leads_followers": 9,
        "instagram_bio_leads": 11,
        "bookings_email": 12,
        "bookings_fb": 13,
        "avg_booking_value": 14,
        "revenue": 17,
        "roas": 18,
        "blended_roas": 19,
    },
    "Feb 2026": {
        "campaign_spend": 5,
        "impressions": 6,
        "profile_visits": 7,
        "leads_followers": 9,
        "instagram_bio_leads": 11,
        "bookings_email": 12,
        "bookings_fb": 13,
        "avg_booking_value": 15,
        "revenue": 17,
        "roas": 18,
        "blended_roas": 19,
    },
    "Jan 2026": {
        "campaign_spend": 4,
        "impressions": 5,
        "profile_visits": 6,
        "leads_followers": 8,
        "instagram_bio_leads": 10,
        "bookings_email": 11,
        "bookings_fb": 12,
        "avg_booking_value": 14,
        "revenue": 16,
        "roas": 17,
        "blended_roas": 18,
    },
}

DISCOVERY_LABELS = {"followers", "new leads", "leads"}
RETARGETING_LABELS = {"retargeting"}
NOISE_TOKENS = {
    "ads",
    "ad",
    "account",
    "meta",
    "new",
}
MANUAL_ALIASES = {
    "flohom ads": "flohom",
    "the cohost company ads": "the cohost company",
    "paradise pointe ads": "paradise pointe",
    "starlight haven hot springs": "starlight haven hot springs",
    "starlight haven weiss lake": "starlight haven weiss lake",
    "stay saluda ads": "stay saluda",
    "asheville river cabins ads": "asheville river cabins",
    "away2pa": "away2pa",
    "awayframes ad account": "awayframes",
    "dwell luxury rentals ads": "dwell luxury rentals",
    "home base bnbs": "home base",
    "myrinn ad account 2 0": "myrinn",
    "pine valley cabins georgia ad account": "pine valley cabins",
    "reflections resort ad": "reflections resorts",
    "roundhouse residences ads": "roundhouse resort spa",
    "stay luxe ads": "stayluxe",
    "stay on 30a ads": "stay on 30a",
    "wauhatchie woodlands timberroot east coast": "wauhatchie woodlands",
    "wanderin star farms new ad account": "wanderin star farms",
    "yosemite meta ads": "yosemite dream stays",
}


def clean_text(value: object) -> str:
    return " ".join(str(value or "").replace("\n", " ").split())


def slugify(value: str) -> str:
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", value.lower()))


def normalize_name(value: str) -> str:
    lowered = clean_text(value).lower()
    lowered = re.sub(r"[^a-z0-9]+", " ", lowered).strip()
    lowered = re.sub(r"\b\d+\b", lambda match: match.group(0) if len(match.group(0)) > 2 else " ", lowered)
    lowered = " ".join(lowered.split())
    return lowered


def generate_name_keys(value: str) -> Iterable[str]:
    base = normalize_name(value)
    if not base:
        return []

    candidates = {base}
    candidates.add(MANUAL_ALIASES.get(base, base))

    tokens = [token for token in base.split() if token not in NOISE_TOKENS]
    if tokens:
        candidates.add(" ".join(tokens))

    stripped = re.sub(r"\b(ad account|meta ads|ads|ad)\b", " ", base)
    stripped = " ".join(stripped.split())
    if stripped:
        candidates.add(stripped)

    return [candidate for candidate in candidates if candidate]


def numeric(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value)
    if not text or text in {"#DIV/0!", "N/A"}:
        return 0.0
    try:
        return float(text.replace(",", "").replace("$", ""))
    except ValueError:
        return 0.0


def nullable_ratio(numerator: float, denominator: float) -> float | None:
    if denominator <= 0:
        return None
    return numerator / denominator


@dataclass
class CampaignTotals:
    spend: float = 0.0
    impressions: float = 0.0
    profile_visits: float = 0.0
    leads_followers: float = 0.0
    instagram_bio_leads: float = 0.0
    bookings_email: float = 0.0
    bookings_fb: float = 0.0
    revenue: float = 0.0
    roas_revenue: float = 0.0
    blended_revenue: float = 0.0

    def merge_row(self, row: dict[str, object]) -> None:
        spend = numeric(row["campaign_spend"])
        revenue = numeric(row["revenue"])
        roas = numeric(row["roas"])
        blended_roas = numeric(row["blended_roas"])

        self.spend += spend
        self.impressions += numeric(row["impressions"])
        self.profile_visits += numeric(row["profile_visits"])
        self.leads_followers += numeric(row["leads_followers"])
        self.instagram_bio_leads += numeric(row["instagram_bio_leads"])
        self.bookings_email += numeric(row["bookings_email"])
        self.bookings_fb += numeric(row["bookings_fb"])
        self.revenue += revenue
        self.roas_revenue += revenue
        self.blended_revenue += blended_roas * spend if blended_roas > 0 else roas * spend

    def to_dict(self, avg_booking_value: float | None) -> dict[str, float | str | None]:
        total_bookings = self.bookings_email + self.bookings_fb
        cost_per_visit = nullable_ratio(self.spend, self.profile_visits)
        cost_per_lead_follower = nullable_ratio(self.spend, self.leads_followers)
        cost_per_booking = nullable_ratio(self.spend, total_bookings)
        pct_avg_booking_value = (
            nullable_ratio(cost_per_booking, avg_booking_value)
            if cost_per_booking is not None and avg_booking_value
            else None
        )

        return {
            "spend": round(self.spend, 2),
            "impressions": round(self.impressions),
            "profileVisits": round(self.profile_visits),
            "costPerVisit": cost_per_visit,
            "leadsFollowers": round(self.leads_followers),
            "costPerLeadFollower": cost_per_lead_follower,
            "igBioLeads": round(self.instagram_bio_leads),
            "bookingsEmail": round(self.bookings_email),
            "bookingsFb": round(self.bookings_fb),
            "avgBookingValue": avg_booking_value,
            "costPerBooking": cost_per_booking,
            "pctAvgBookingValue": pct_avg_booking_value,
            "revenue": round(self.revenue, 2),
            "roas": nullable_ratio(self.roas_revenue, self.spend),
            "blendedRoas": nullable_ratio(self.blended_revenue, self.spend),
        }


@dataclass
class ClientBlock:
    raw_name: str
    normalized_name: str | None = None
    slug: str | None = None
    avg_booking_value: float | None = None
    discovery: CampaignTotals = field(default_factory=CampaignTotals)
    retargeting: CampaignTotals = field(default_factory=CampaignTotals)


def load_client_lookup() -> dict[str, dict[str, str]]:
    payload = json.loads(PERFORMANCE_JSON_PATH.read_text(encoding="utf-8"))
    lookup: dict[str, dict[str, str]] = {}
    for client in payload["clients"]:
        name = client["name"]
        normalized = normalize_name(name)
        lookup[normalized] = client
    return lookup


def resolve_client(raw_name: str, lookup: dict[str, dict[str, str]]) -> tuple[str | None, str | None, str | None]:
    keys = list(generate_name_keys(raw_name))
    for key in keys:
        client = lookup.get(key)
        if client:
            return key, client["slug"], client["name"]

    all_keys = list(lookup.keys())
    best_key = difflib.get_close_matches(normalize_name(raw_name), all_keys, n=1, cutoff=0.75)
    if best_key:
        client = lookup[best_key[0]]
        return best_key[0], client["slug"], client["name"]

    return None, None, None


def classify_campaign(value: object) -> str | None:
    label = clean_text(value).lower()
    if label in DISCOVERY_LABELS:
        return "Discovery"
    if label in RETARGETING_LABELS:
        return "Retargeting"
    return None


def parse_sheet(sheet_name: str) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    lookup = load_client_lookup()
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    columns = SHEET_COLUMN_MAPS[sheet_name]

    parsed_blocks: list[ClientBlock] = []
    unresolved: list[dict[str, object]] = []
    current: ClientBlock | None = None

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=20, values_only=True), start=2):
        raw_name = clean_text(row[0])
        if raw_name:
            resolved_key, resolved_slug, resolved_name = resolve_client(raw_name, lookup)
            current = ClientBlock(
                raw_name=raw_name,
                normalized_name=resolved_name,
                slug=resolved_slug,
            )
            parsed_blocks.append(current)
            if not resolved_slug:
                unresolved.append({"row": row_index, "rawName": raw_name})

        if current is None:
            continue

        campaign = classify_campaign(row[3])
        if campaign is None:
            continue

        avg_booking_value = numeric(row[columns["avg_booking_value"]])
        if avg_booking_value > 0:
            current.avg_booking_value = avg_booking_value

        row_payload = {
            "campaign_spend": row[columns["campaign_spend"]],
            "impressions": row[columns["impressions"]],
            "profile_visits": row[columns["profile_visits"]],
            "leads_followers": row[columns["leads_followers"]],
            "instagram_bio_leads": row[columns["instagram_bio_leads"]],
            "bookings_email": row[columns["bookings_email"]],
            "bookings_fb": row[columns["bookings_fb"]],
            "revenue": row[columns["revenue"]],
            "roas": row[columns["roas"]],
            "blended_roas": row[columns["blended_roas"]],
        }

        if campaign == "Discovery":
            current.discovery.merge_row(row_payload)
        else:
            current.retargeting.merge_row(row_payload)

    rows: list[dict[str, object]] = []
    for block in parsed_blocks:
        avg_booking_value = block.avg_booking_value
        for campaign_type, totals in [("Discovery", block.discovery), ("Retargeting", block.retargeting)]:
            if totals.spend <= 0 and totals.impressions <= 0 and totals.profile_visits <= 0:
                continue
            row = {
                "rawName": block.raw_name,
                "matchedName": block.normalized_name,
                "slug": block.slug,
                "campaignType": campaign_type,
                "month": sheet_name,
            }
            row.update(totals.to_dict(avg_booking_value))
            rows.append(row)

    return rows, unresolved


def main() -> None:
    summaries = []
    for sheet_name in SHEET_COLUMN_MAPS:
        rows, unresolved = parse_sheet(sheet_name)
        summaries.append({
            "sheet": sheet_name,
            "rowCount": len(rows),
            "unresolvedClients": unresolved,
            "sample": rows[:8],
            "inspiredRetreats": [row for row in rows if row["slug"] == "inspired-retreats"],
        })

    summary = {"sheets": summaries}
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
